




import streamlit as st
import pandas as pd
import numpy as np

from modules.calculation import calculations
from modules.plotting import plotting
from modules import data_loader

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

from pathlib import Path
import os

# Define the base directory relative to the current script file
# Path(__file__).resolve().parent gets the directory containing the script being run
BASE_DIR = Path(__file__).resolve().parent / "sec_wise_futures_data"

st.set_page_config(layout="wide", page_title="Pro Quant Pair Studio")
# st.set_page_config(layout="wide", page_title="Pro Quant Pair Studio)

# 1. MODIFIED: Added lookback and std_mult arguments to signature
def dataframe_to_formatted_excel(df, lookback, std_mult):
    
    df_processed = df.copy() 

    stock_a = str((df_processed['SYMBOL_A']).unique()[0])
    stock_b = str((df_processed['SYMBOL_B']).unique()[0])

    # Try/Except block added to handle dataframes that might not have all columns (like normalized data)
    try:
        df_processed.drop(['INSTRUMENT_A','SYMBOL_A','OPEN_INT*_A','TRD_VAL_A','TRD_QTY_A','NO_OF_CONT_A','NO_OF_TRADE_A',
                            'TIMESTAMP_B','INSTRUMENT_B','SYMBOL_B','OPEN_INT*_B','TRD_VAL_B','TRD_QTY_B','NO_OF_CONT_B','NO_OF_TRADE_B',
                            'NEXT_EXP_DATE_A','NEXT_EXP_DATE_B'], axis=1, inplace=True)
    except KeyError:
        pass # Ignore if columns don't exist
    
    if 'TIMESTAMP_A' in df_processed.columns:
        df_processed.rename(columns={'TIMESTAMP_A': 'Date'}, inplace=True)

    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:

        df_processed.to_excel(writer, sheet_name="Pair Data", index=False, startrow=1) 
        
        wb = writer.book
        ws = writer.sheets["Pair Data"]

        ws.insert_cols(14) # Gap 3
        ws.insert_cols(9)  # Gap 2
        ws.insert_cols(2)  # Gap 1
        
        # --- Stock A Header ---
        ws.merge_cells("C1:H1") 
        ws["C1"] = stock_a
        ws["C1"].font = Font(bold=True)
        ws["C1"].alignment = Alignment(horizontal="center")

        # --- Stock B Header ---
        ws.merge_cells("J1:O1")
        ws["J1"] = stock_b
        ws["J1"].font = Font(bold=True)
        ws["J1"].alignment = Alignment(horizontal="center")

        # --- MODIFIED: Indicators Header with Parameters ---
        ws.merge_cells("Q1:W1")
        # Creating the detailed string requested
        header_text = f"INDICATORS (RSI/Z-Score: {lookback}D | Bands: ±{std_mult} SD)"
        ws["Q1"] = header_text
        ws["Q1"].font = Font(bold=True)
        ws["Q1"].alignment = Alignment(horizontal="center")

        for cell in ws[2]: 
            if cell.value: 
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_length = 0
            column = col[1].column_letter 
            for cell in col:
                try: 
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length) * 1.2
            print(column + str(adjusted_width))
            ws.column_dimensions[column].width = max(12,adjusted_width)

    buffer.seek(0)
    return buffer


def preprocess_futures(df):
    df = df.copy()
    df.drop(['OPEN_INT_A', 'OPEN_INT_B'], axis=1, inplace=True, errors='ignore')

    for leg in ['A', 'B']:
        date_col = f'TIMESTAMP_{leg}'
        expiry_col = f'EXP_DATE_{leg}'
        symbol_col = f'SYMBOL_{leg}'
        gap_col = f'expiry_gap_{leg}'

        df[date_col] = pd.to_datetime(df[date_col])
        df[expiry_col] = pd.to_datetime(df[expiry_col])

        df = df.sort_values([symbol_col, expiry_col, date_col])

        next_exp_col = f'NEXT_EXP_DATE_{leg}'
        df[next_exp_col] = (df.groupby(symbol_col)[expiry_col].shift(-1))

        df[gap_col] = np.where(
            df[date_col] < df[expiry_col],
            (df[expiry_col] - df[date_col]).dt.days,
            (df[next_exp_col] - df[date_col]).dt.days
        )
        df.loc[df[next_exp_col].isna(), gap_col] = np.nan

    return df


# 2. MODIFIED: Added lookback and std_mult arguments to signature
def render_chart_with_csv(fig, df_to_download, lookback, std_mult, filename_prefix="data"):
    st.plotly_chart(fig, use_container_width=True)

    # ---- CSV ----
    # st.download_button(
    #     "📥 Download CSV",
    #     df_to_download.to_csv(index=False).encode("utf-8"),
    #     file_name=f"{filename_prefix}.csv",
    #     mime="text/csv"
    # )

    # ---- EXCEL (FORMATTED) ----
    # Passing the parameters to the excel function
    excel_file = dataframe_to_formatted_excel(df_to_download, lookback, std_mult)

    st.download_button(
        "📊 Download Excel (Formatted)",
        data=excel_file,
        file_name=f"{filename_prefix}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


st.title("⚡ Pro Quant Pairs Studio")
# st.title(f"{BASE_DIR}")

# --- SIDEBAR ---
st.sidebar.header("1. Select Pair")

sectors = data_loader.get_sectors(BASE_DIR)
selected_sector = st.sidebar.selectbox("Select Sector", sectors)

if selected_sector:
    stocks = data_loader.get_stocks(BASE_DIR, selected_sector)

    col1, col2 = st.sidebar.columns(2)
    with col1:
        stock_a = st.selectbox("Stock A (Long)", stocks, index=0)
    with col2:
        stock_b = st.selectbox(
            "Stock B (Short)",
            stocks,
            index=1 if len(stocks) > 1 else 0
        )

    st.sidebar.divider()
    st.sidebar.header("2. Parameters")

    lookback = st.sidebar.number_input("Lookback (SMA)", 5, 200, 20)
    std_dev_mult = st.sidebar.slider("Std Dev Multiplier", 0.5, 4.0, 2.0, 0.1)
    norm_start_date = st.sidebar.date_input(
        "Performance Start Date",
        pd.to_datetime("2023-01-01")
    )

    if stock_a != stock_b:
        with st.spinner("Crunching Data..."):
            raw_df = data_loader.load_pair_data(
                BASE_DIR, selected_sector, stock_a, stock_b
            )

        if raw_df is not None and not raw_df.empty:

            raw_df = preprocess_futures(raw_df)

            processed_df = calculations.calculate_pair_metrics(
                raw_df,
                window=lookback,
                std_dev_multiplier=std_dev_mult
            )

            print("Processed DF Head:")
            print(processed_df.head(10))

            st.subheader("1. Price Ratio & Bollinger Bands")
            fig_ratio, data_ratio = plotting.plot_pair_ratio(
                processed_df, stock_a, stock_b, std_dev_mult
            )
            # 3. MODIFIED: Passing lookback and std_dev_mult
            render_chart_with_csv(fig_ratio, data_ratio, lookback, std_dev_mult, "bollinger_ratio_data")

            st.divider()

            st.subheader("2. Normalized Performance Comparison")
            norm_df = calculations.calculate_normalized_performance(
                raw_df, norm_start_date
            )

            if not norm_df.empty:
                dom_stats = calculations.calculate_dominance_stats(norm_df)

                if dom_stats:
                    st.markdown("##### 🏆 Performance Dominance")
                    c1, c2, c3, c4 = st.columns(4)
                    c1.metric(stock_a, f"{dom_stats['hist_a']:.1f}%")
                    c2.metric(stock_b, f"{dom_stats['hist_b']:.1f}%")
                    c3.metric(f"{stock_a} (3M)", f"{dom_stats['rec_a']:.1f}%")
                    c4.metric(f"{stock_b} (3M)", f"{dom_stats['rec_b']:.1f}%")

                fig_norm, data_norm = plotting.plot_normalized_comparison(
                    norm_df, stock_a, stock_b, str(norm_start_date)
                )
                # 3. MODIFIED: Passing lookback and std_dev_mult
                render_chart_with_csv(
                    fig_norm, data_norm, lookback, std_dev_mult, "normalized_performance_data"
                )
            else:
                st.warning("No data available after selected date.")

            st.divider()


            st.subheader("3. RSI Momentum Indicator")
            fig_rsi, data_rsi = plotting.plot_secondary_indicators(
                processed_df, std_dev_mult
            )
            # 3. MODIFIED: Passing lookback and std_dev_mult
            render_chart_with_csv(fig_rsi, data_rsi, lookback, std_dev_mult, "rsi_data")

            st.divider()


            st.subheader("4. Statistical Distribution")
            fig_dist = plotting.plot_zscore_distribution(processed_df)
            st.plotly_chart(fig_dist, use_container_width=True)

            st.divider()


            st.subheader("5. Quarterly Statistical Breakdown")
            fig_step = plotting.plot_quarterly_step_chart(processed_df)
            st.plotly_chart(fig_step, use_container_width=True)

        else:
            st.error("Not enough overlapping data.")
    else:
        st.warning("Select two different stocks.")
else:
    st.info("👈 Select a sector to begin.")