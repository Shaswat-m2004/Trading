


# # import streamlit as st
# # import pandas as pd
# # import numpy as np
# # from io import BytesIO
# # from openpyxl import Workbook
# # from openpyxl.styles import Font, Alignment
# # import concurrent.futures # For Parallel Processing

# # # Custom Modules
# # from modules.calculation import calculations
# # from modules.plotting import plotting
# # from modules import data_loader

# # # --- HELPERS (Excel & Preprocess) ---
# # # (Keeping excel logic same)
# # def dataframe_to_formatted_excel(df, lookback, std_mult):
# #     df_processed = df.copy()
# #     stock_a = str((df_processed['SYMBOL_A']).unique()[0])
# #     stock_b = str((df_processed['SYMBOL_B']).unique()[0])
    
# #     try:
# #         df_processed.drop(['INSTRUMENT_A','SYMBOL_A','OPEN_INT*_A','TRD_VAL_A','TRD_QTY_A','NO_OF_CONT_A','NO_OF_TRADE_A',
# #                            'TIMESTAMP_B','INSTRUMENT_B','SYMBOL_B','OPEN_INT*_B','TRD_VAL_B','TRD_QTY_B','NO_OF_CONT_B','NO_OF_TRADE_B',
# #                            'NEXT_EXP_DATE_A','NEXT_EXP_DATE_B'], axis=1, inplace=True)
# #     except KeyError: pass
    
# #     if 'TIMESTAMP_A' in df_processed.columns: 
# #         df_processed.rename(columns={'TIMESTAMP_A': 'Date'}, inplace=True)

# #     buffer = BytesIO()
# #     with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
# #         df_processed.to_excel(writer, sheet_name="Pair Data", index=False, startrow=1)
# #         wb = writer.book
# #         ws = writer.sheets["Pair Data"]
        
# #         ws.insert_cols(14); ws.insert_cols(9); ws.insert_cols(2)
        
# #         # Headers
# #         ws.merge_cells("C1:H1"); ws["C1"] = stock_a
# #         ws["C1"].font = Font(bold=True); ws["C1"].alignment = Alignment(horizontal="center")
        
# #         ws.merge_cells("J1:O1"); ws["J1"] = stock_b
# #         ws["J1"].font = Font(bold=True); ws["J1"].alignment = Alignment(horizontal="center")
        
# #         ws.merge_cells("Q1:W1")
# #         ws["Q1"] = f"INDICATORS (RSI/Z-Score: {lookback}D | Bands: ±{std_mult} SD)"
# #         ws["Q1"].font = Font(bold=True); ws["Q1"].alignment = Alignment(horizontal="center")

# #         # Auto-width
# #         for col in ws.columns:
# #             max_length = 0
# #             column = col[1].column_letter 
# #             for cell in col:
# #                 try: 
# #                     if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
# #                 except: pass
# #             ws.column_dimensions[column].width = max(12, (max_length * 1.2))

# #     buffer.seek(0)
# #     return buffer

# # def preprocess_futures(df):
# #     df = df.copy()
# #     df.drop(['OPEN_INT_A', 'OPEN_INT_B'], axis=1, inplace=True, errors='ignore')
# #     for leg in ['A', 'B']:
# #         date_col = f'TIMESTAMP_{leg}'; expiry_col = f'EXP_DATE_{leg}'; symbol_col = f'SYMBOL_{leg}'; gap_col = f'expiry_gap_{leg}'
# #         df[date_col] = pd.to_datetime(df[date_col]); df[expiry_col] = pd.to_datetime(df[expiry_col])
# #         df = df.sort_values([symbol_col, expiry_col, date_col])
# #         next_exp_col = f'NEXT_EXP_DATE_{leg}'
# #         df[next_exp_col] = (df.groupby(symbol_col)[expiry_col].shift(-1))
# #         df[gap_col] = np.where(df[date_col] < df[expiry_col], (df[expiry_col] - df[date_col]).dt.days, (df[next_exp_col] - df[date_col]).dt.days)
# #         df.loc[df[next_exp_col].isna(), gap_col] = np.nan
# #     return df

# # def render_chart_with_csv(fig, df_to_download, lookback, std_mult, filename_prefix="data"):
# #     st.plotly_chart(fig, use_container_width=True)
# #     excel_file = dataframe_to_formatted_excel(df_to_download, lookback, std_mult)
# #     st.download_button("📊 Download Excel", data=excel_file, file_name=f"{filename_prefix}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=filename_prefix)

# # def render_csv(fig, df_to_download, lookback, std_mult, filename_prefix="data"):
# #     st.plotly_chart(fig, use_container_width=True)
# #     file = df_to_download.to_csv(index=False).encode('utf-8')
# #     st.download_button("📊 Download CSV", data=file, file_name=f"{filename_prefix}.csv", mime="text/csv", key=filename_prefix)



# # # --- OPTIMIZATION 1: Caching Data Loading & Calculation ---
# # # Ye function data load karega aur calculations karega.
# # # Jab tak parameters same hain, ye dobara run nahi hoga -> Instant Speed ⚡
# # @st.cache_data(show_spinner=False)
# # def get_cached_data_and_calc(base_dir, sector, stock_a, stock_b, lookback, std_dev_mult, norm_start_date):
# #     # 1. Load
# #     raw_df = data_loader.load_pair_data(base_dir, sector, stock_a, stock_b)
# #     if raw_df is None or raw_df.empty:
# #         return None, None, None, None

# #     # 2. Preprocess
# #     raw_df = preprocess_futures(raw_df)

# #     # 3. Calculate
# #     processed_df = calculations.calculate_pair_metrics(raw_df, window=lookback, std_dev_multiplier=std_dev_mult)
# #     processed_df = calculations.calculate_quarterly_structure(processed_df)
# #     processed_df_for_semiannual = calculations.calculate_semiannual_weekly_structure(processed_df)
# #     stats = calculations.calculate_bollinger_reversion_stats(processed_df)
# #     print(processed_df_for_semiannual.columns)
    
# #     # 4. Normalize Data
# #     norm_df = calculations.calculate_normalized_performance(raw_df, norm_start_date)
    
# #     return processed_df, stats, norm_df, raw_df,processed_df_for_semiannual

# # # --- OPTIMIZATION 2: Parallel Plot Generation ---
# # def generate_all_plots_parallel(processed_df, norm_df, stock_a, stock_b, std_dev_mult, norm_start_date_str,processed_df_for_semiannual):
# #     with concurrent.futures.ThreadPoolExecutor() as executor:
# #         # Submit tasks
# #         future_ratio = executor.submit(plotting.plot_pair_ratio, processed_df, stock_a, stock_b, std_dev_mult)
        
# #         future_norm = None
# #         if not norm_df.empty:
# #             future_norm = executor.submit(plotting.plot_normalized_comparison, norm_df, stock_a, stock_b, norm_start_date_str)
            
# #         future_rsi = executor.submit(plotting.plot_secondary_indicators, processed_df, std_dev_mult)
# #         future_dist = executor.submit(plotting.plot_zscore_distribution, processed_df)
# #         future_structure = executor.submit(plotting.plot_quarterly_structure, processed_df, stock_a, stock_b)
# #         future_annual = executor.submit(plotting.plot_semiannual_weekly_structure,processed_df_for_semiannual,stock_a,stock_b)
# #         res_ratio = future_ratio.result()
# #         # ...
# #         res_structure = future_structure.result()
# #         # future_step = executor.submit(plotting.plot_quarterly_step_chart, processed_df)
# #         # future_semiannual = executor.submit(plotting.plot_semiannual_weekly_line, processed_df)

# #         # Collect results
# #         res_ratio = future_ratio.result()
# #         res_norm = future_norm.result() if future_norm else (None, None)
# #         res_rsi = future_rsi.result()
# #         res_dist = future_dist.result()
# #         res_annual = future_annual.result()
# #         # res_step = future_step.result()
# #         # res_semiannual = future_semiannual.result()

# #         print("hello................................................................")
# #         print("-"*20)
# #         print(res_annual)
        
# #     # return res_ratio, res_norm, res_rsi, res_dist, res_step, res_semiannual
# #     return res_ratio, res_norm, res_rsi, res_dist,res_structure,res_annual


# # # --- MAIN RUN FUNCTION ---
# # def run(BASE_DIR, selected_sector, stock_a, stock_b, lookback, std_dev_mult, norm_start_date):
    
# #     st.header("📈 Pro Quant Chart Studio")
    
# #     if not selected_sector:
# #         st.info("👈 Please select a sector from the sidebar to begin.")
# #         return

# #     st.markdown(f"**Analyzing:** {stock_a} vs {stock_b} | **Sector:** {selected_sector}")

# #     if stock_a != stock_b:
# #         # --- STEP 1: Fast Data & Calculation (Cached) ---
# #         with st.spinner("⚡ Crunching Numbers..."):
# #             processed_df, stats, norm_df, raw_df,processed_df_for_semiannual = get_cached_data_and_calc(
# #                 BASE_DIR, selected_sector, stock_a, stock_b, lookback, std_dev_mult, norm_start_date
# #             )

# #         if processed_df is not None:
            
# #             # --- STEP 2: Metrics (Instant Display) ---
# #             m1, m2, m3, m4 = st.columns(4)
# #             m1.metric("-2σ Touches", stats["touches"])
# #             m2.metric("Reversions", f"{stats['mean_reversions']} ({stats['success_rate']:.1f}%)")
# #             m3.metric("Failures", stats["failures"])
# #             m4.metric("Avg Days", f"{stats['avg_days_to_mean']:.1f}")

# #             (fig_ratio, data_ratio), (fig_norm, data_norm), (fig_rsi, data_rsi), fig_dist,(fig_struct, data_struct),(fig, plot_df) = generate_all_plots_parallel(
# #                 processed_df, norm_df, stock_a, stock_b, std_dev_mult, str(norm_start_date),processed_df_for_semiannual
# #             )

# #             st.subheader("1. Price Ratio & Bollinger Bands")
# #             render_chart_with_csv(fig_ratio, data_ratio, lookback, std_dev_mult, "bollinger_ratio_data")
            
# #             st.divider()

# #             # 2. Normalized Chart
# #             st.subheader("2. Normalized Performance")
# #             if fig_norm:
# #                 render_chart_with_csv(fig_norm, data_norm, lookback, std_dev_mult, "normalized_data")
# #             else:
# #                 st.warning("No data found for Normalized Chart after start date.")
            
# #             st.divider()

# #             # 3. RSI Chart
# #             st.subheader("3. RSI Momentum")
# #             render_chart_with_csv(fig_rsi, data_rsi, lookback, std_dev_mult, "rsi_data")
            
# #             st.divider()
# #             st.subheader("4. Z-Score Dist")
# #             st.plotly_chart(fig_dist, use_container_width=True)
    
# #             st.divider()
    
# #             # --- NEW SECTION ---
# #             st.subheader("4. Quarterly Structure Analysis")
# #             st.markdown("""
# #             *Visualizes the structural range for every 3 months based on synthetic OHLC Ratio.*
# #             - **Green Line:** High Band `(Max Close + Max High) / 2`
# #             - **Red Line:** Low Band `(Min Close + Min Low) / 2`
# #             """)
# #             render_chart_with_csv(fig_struct, data_struct, lookback, std_dev_mult, "quarterly_structure_data")
            
# #             st.divider()
            
# #             # --- NEW SECTION ---
# #             st.subheader("4. Quarterly Structure Analysis")
# #             st.markdown("""
# #             *Visualizes the structural range for every 3 months based on synthetic OHLC Ratio.*
# #             - **Green Line:** High Band `(Max Close + Max High) / 2`
# #             - **Red Line:** Low Band `(Min Close + Min Low) / 2`
# #             """)
# #             render_csv(fig, plot_df, lookback, std_dev_mult, "annualy_structure_data")
            
# #             st.divider()
# #         else: 
# #             st.error("Not enough overlapping data found for this pair.")
# #     else: 
# #         st.warning("⚠️ Please select two **different** stocks in the sidebar.")








# import streamlit as st
# import pandas as pd
# import numpy as np
# from io import BytesIO
# from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment
# import concurrent.futures # For Parallel Processing

# # Custom Modules
# from modules.calculation import calculations
# from modules.plotting import plotting
# from modules import data_loader

# # --- HELPERS (Excel & Preprocess) ---
# # ... (Keep your existing dataframe_to_formatted_excel and preprocess_futures functions AS IS) ...
# # Paste dataframe_to_formatted_excel here (same as your code)
# def dataframe_to_formatted_excel(df, lookback, std_mult):
#     df_processed = df.copy()
#     stock_a = str((df_processed['SYMBOL_A']).unique()[0])
#     stock_b = str((df_processed['SYMBOL_B']).unique()[0])
    
#     try:
#         df_processed.drop(['INSTRUMENT_A','SYMBOL_A','OPEN_INT*_A','TRD_VAL_A','TRD_QTY_A','NO_OF_CONT_A','NO_OF_TRADE_A',
#                            'TIMESTAMP_B','INSTRUMENT_B','SYMBOL_B','OPEN_INT*_B','TRD_VAL_B','TRD_QTY_B','NO_OF_CONT_B','NO_OF_TRADE_B',
#                            'NEXT_EXP_DATE_A','NEXT_EXP_DATE_B'], axis=1, inplace=True)
#     except KeyError: pass
    
#     if 'TIMESTAMP_A' in df_processed.columns: 
#         df_processed.rename(columns={'TIMESTAMP_A': 'Date'}, inplace=True)

#     buffer = BytesIO()
#     with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
#         df_processed.to_excel(writer, sheet_name="Pair Data", index=False, startrow=1)
#         wb = writer.book
#         ws = writer.sheets["Pair Data"]
        
#         ws.insert_cols(14); ws.insert_cols(9); ws.insert_cols(2)
        
#         # Headers
#         ws.merge_cells("C1:H1"); ws["C1"] = stock_a
#         ws["C1"].font = Font(bold=True); ws["C1"].alignment = Alignment(horizontal="center")
        
#         ws.merge_cells("J1:O1"); ws["J1"] = stock_b
#         ws["J1"].font = Font(bold=True); ws["J1"].alignment = Alignment(horizontal="center")
        
#         ws.merge_cells("Q1:W1")
#         ws["Q1"] = f"INDICATORS (RSI/Z-Score: {lookback}D | Bands: ±{std_mult} SD)"
#         ws["Q1"].font = Font(bold=True); ws["Q1"].alignment = Alignment(horizontal="center")

#         # Auto-width
#         for col in ws.columns:
#             max_length = 0
#             column = col[1].column_letter 
#             for cell in col:
#                 try: 
#                     if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
#                 except: pass
#             ws.column_dimensions[column].width = max(12, (max_length * 1.2))

#     buffer.seek(0)
#     return buffer

# def preprocess_futures(df):
#     df = df.copy()
#     df.drop(['OPEN_INT_A', 'OPEN_INT_B'], axis=1, inplace=True, errors='ignore')
#     for leg in ['A', 'B']:
#         date_col = f'TIMESTAMP_{leg}'; expiry_col = f'EXP_DATE_{leg}'; symbol_col = f'SYMBOL_{leg}'; gap_col = f'expiry_gap_{leg}'
#         df[date_col] = pd.to_datetime(df[date_col]); df[expiry_col] = pd.to_datetime(df[expiry_col])
#         df = df.sort_values([symbol_col, expiry_col, date_col])
#         next_exp_col = f'NEXT_EXP_DATE_{leg}'
#         df[next_exp_col] = (df.groupby(symbol_col)[expiry_col].shift(-1))
#         df[gap_col] = np.where(df[date_col] < df[expiry_col], (df[expiry_col] - df[date_col]).dt.days, (df[next_exp_col] - df[date_col]).dt.days)
#         df.loc[df[next_exp_col].isna(), gap_col] = np.nan
#     return df

# # --- MODIFIED RENDERING FUNCTIONS ---
# # Instead of generating file, we register a task and show a placeholder

# def render_chart_deferred(download_queue, fig, df_to_download, lookback, std_mult, filename_prefix, file_type="excel"):
#     """
#     Renders chart immediately. Creates a placeholder for the button 
#     and adds the data task to the queue.
#     """
#     # 1. Show Chart Immediately
#     st.plotly_chart(fig, use_container_width=True)
    
#     # 2. Create Placeholder
#     placeholder = st.empty()
#     placeholder.caption(f"⏳ Preparing {file_type.upper()}...")

#     # 3. Add to Queue for processing later
#     download_queue.append({
#         "placeholder": placeholder,
#         "df": df_to_download,
#         "lookback": lookback,
#         "std_mult": std_mult,
#         "filename": filename_prefix,
#         "type": file_type
#     })

# # --- DATA CACHING ---
# @st.cache_data(show_spinner=False)
# def get_cached_data_and_calc(base_dir, sector, stock_a, stock_b, lookback, std_dev_mult, norm_start_date):
#     # 1. Load
#     raw_df = data_loader.load_pair_data(base_dir, sector, stock_a, stock_b)
#     if raw_df is None or raw_df.empty:
#         return None, None, None, None, None

#     # 2. Preprocess
#     raw_df = preprocess_futures(raw_df)



#     # 3. Calculate
#     processed_df = calculations.calculate_pair_metrics(raw_df, window=lookback, std_dev_multiplier=std_dev_mult)
#     processed_df = calculations.calculate_quarterly_structure(processed_df)
#     processed_df_for_semiannual = calculations.calculate_semiannual_weekly_structure(processed_df)
#     stats = calculations.calculate_bollinger_reversion_stats(processed_df)
#     spread = calculations.get_spread_series(processed_df)

#     spread = spread.tail(250)

#     relationship_metrics = calculations.calculate_relationship_metrics(processed_df)
 

#     cointegration_stats = calculations.calculate_cointegration_metrics(spread)
#     hurst_value = calculations.calculate_hurst_exponent(spread)
    
#     # 4. Normalize Data
#     norm_df = calculations.calculate_normalized_performance(raw_df, norm_start_date)
    
#     # return processed_df, stats, norm_df, raw_df, processed_df_for_semiannual
#     return (
#     processed_df,
#     stats,
#     norm_df,
#     raw_df,
#     processed_df_for_semiannual,
#     cointegration_stats,
#     hurst_value,
#     relationship_metrics

#     )


# # --- PLOTTING ---
# def generate_all_plots_parallel(processed_df, norm_df, stock_a, stock_b, std_dev_mult, norm_start_date_str, processed_df_for_semiannual):
#     with concurrent.futures.ThreadPoolExecutor() as executor:
#         future_ratio = executor.submit(plotting.plot_pair_ratio, processed_df, stock_a, stock_b, std_dev_mult)
        
#         future_norm = None
#         if not norm_df.empty:
#             future_norm = executor.submit(plotting.plot_normalized_comparison, norm_df, stock_a, stock_b, norm_start_date_str)
            
#         future_rsi = executor.submit(plotting.plot_secondary_indicators, processed_df, std_dev_mult)
#         future_dist = executor.submit(plotting.plot_zscore_distribution, processed_df)
#         future_structure = executor.submit(plotting.plot_quarterly_structure, processed_df, stock_a, stock_b)
#         future_annual = executor.submit(plotting.plot_semiannual_weekly_structure, processed_df_for_semiannual, stock_a, stock_b)

#         res_ratio = future_ratio.result()
#         res_norm = future_norm.result() if future_norm else (None, None)
#         res_rsi = future_rsi.result()
#         res_dist = future_dist.result()
#         res_structure = future_structure.result()
#         res_annual = future_annual.result()
        
#     return res_ratio, res_norm, res_rsi, res_dist, res_structure, res_annual


# # --- MAIN RUN FUNCTION ---
# def run(BASE_DIR, selected_sector, stock_a, stock_b, lookback, std_dev_mult, norm_start_date):
    
#     st.header("📈 Pro Quant Chart Studio")
    
#     if not selected_sector:
#         st.info("👈 Please select a sector from the sidebar to begin.")
#         return

#     st.markdown(f"**Analyzing:** {stock_a} vs {stock_b} | **Sector:** {selected_sector}")

#     if stock_a != stock_b:
#         # --- STEP 1: Fast Data & Calculation (Cached) ---
#         with st.spinner("⚡ Crunching Numbers..."):
#             processed_df, stats, norm_df, raw_df, processed_df_for_semiannual,coint_stats, hurst,rel_metrics = get_cached_data_and_calc(
#                 BASE_DIR, selected_sector, stock_a, stock_b, lookback, std_dev_mult, norm_start_date
#             )

#         if processed_df is not None:
            
#             # --- STEP 2: Metrics ---
#             m1, m2, m3, m4 = st.columns(4)
#             m1.metric("-2σ Touches", stats["touches"])
#             m2.metric("Reversions", f"{stats['mean_reversions']} ({stats['success_rate']:.1f}%)")
#             m3.metric("Failures", stats["failures"])
#             m4.metric("Avg Days", f"{stats['avg_days_to_mean']:.1f}")
            
#             # Generate Plots
#             (fig_ratio, data_ratio), (fig_norm, data_norm), (fig_rsi, data_rsi), fig_dist, (fig_struct, data_struct), (fig_annual, plot_df_annual) = generate_all_plots_parallel(
#                 processed_df, norm_df, stock_a, stock_b, std_dev_mult, str(norm_start_date), processed_df_for_semiannual
#             )

#             # --- STEP 3: Render Charts & Queue Downloads ---
#             # List to hold tasks for deferred processing
#             download_queue = []

#             st.subheader("1. Price Ratio & Bollinger Bands")
#             # Pass download_queue to helper function
#             render_chart_deferred(download_queue, fig_ratio, data_ratio, lookback, std_dev_mult, "bollinger_ratio_data", "excel")
            
#             st.divider()

#             st.subheader("📊 Pair Relationship Metrics")

#             r1, r2, r3 = st.columns(3)

#             r1.metric(
#                 "Correlation",
#                 f"{rel_metrics['correlation']:.2f}",
#                 delta="Strong" if rel_metrics['correlation'] > 0.6 else "Weak"
#             )

#             r2.metric(
#                 "Covariance",
#                 f"{rel_metrics['covariance']:.2f}"
#             )

#             spread_var = rel_metrics['spread_variance']

#             r3.metric(
#                 "Spread Variance",
#                 f"{spread_var:.4f}",
#                 delta="Stable" if spread_var < spread_var * 1.5 else "Volatile"
#             )

#             st.subheader("📊 Pair Stability & Mean Reversion")

#             c1, c2, c3, c4 = st.columns(4)

#             # Cointegration
#             c1.metric(
#             "Cointegration",
#             "YES" if coint_stats["cointegrated"] else "NO",
#             delta=f"Strength: {coint_stats['strength']:.2f}"
#             )

#             # ADF p-value
#             c2.metric(
#                 "ADF p-value",
#                 f"{coint_stats['p_value']:.4f}",
#             )

#             # ADF Statistic
#             c3.metric(
#                 "ADF Statistic",
#                 f"{coint_stats['adf_stat']:.2f}"
#             )

#             # Hurst Exponent
#             c4.metric(
#                 "Hurst Exponent",
#                 f"{hurst:.2f}",
#                 delta="Mean Reverting" if hurst < 0.5 else "Trending"
#             )


#             st.subheader("2. Normalized Performance")
#             if fig_norm:
#                 render_chart_deferred(download_queue, fig_norm, data_norm, lookback, std_dev_mult, "normalized_data", "excel")
#             else:
#                 st.warning("No data found for Normalized Chart.")
            
#             st.divider()

#             st.subheader("3. RSI Momentum")
#             render_chart_deferred(download_queue, fig_rsi, data_rsi, lookback, std_dev_mult, "rsi_data", "excel")
            
#             st.divider()
            
#             # st.subheader("4. Z-Score Dist")
#             # st.plotly_chart(fig_dist, use_container_width=True)
#             # No download for distribution usually needed, or add if you want
#             # App.py mein jahan plot render kar rahe ho:

#             st.subheader("4. Z-Score Regime Analysis")
#             dist_stats = calculations.calculate_distribution_stats(processed_df)

#             # 2 Columns banayenge: Left mein Chart, Right mein Data
#             col_graph, col_stats = st.columns([3, 1]) 

#             with col_graph:
#                 st.plotly_chart(fig_dist, use_container_width=True)

#             with col_stats:
#                 st.markdown("#### 📊 Sigma Counts")
#                 # Wo function jo maine pichle message mein diya tha (calculate_distribution_stats)
#                 if dist_stats is not None:
#                     # Table ko clean dikhane ke liye
#                     st.dataframe(
#                         dist_stats[['Range', 'Actual_%', 'Theoretical_%']].style.format({
#                             'Actual_%': '{:.1f}%',
#                             'Theoretical_%': '{:.1f}%'
#                         }).background_gradient(subset=['Actual_%'], cmap='Greens'),
#                         height=200,
#                         use_container_width=True,
#                         hide_index=True
#                     )
        
#         # Quick Insight
#         outlier_pct = dist_stats.iloc[3]['Actual_%']
#         if outlier_pct > 1.0:
#             st.error(f"⚠️ High Risk: {outlier_pct:.1f}% data in outliers!")
#         else:
#             st.success("✅ Stable Distribution")
    
#             st.divider()
    
#             st.subheader("5. Quarterly Structure")
#             render_chart_deferred(download_queue, fig_struct, data_struct, lookback, std_dev_mult, "quarterly_structure", "excel")
            
#             st.divider()
            
#             st.subheader("6. Semiannual Structure")
#             # Note: Using CSV for this one as per your original code request
#             render_chart_deferred(download_queue, fig_annual, plot_df_annual, lookback, std_dev_mult, "annualy_structure", "csv")
            
#             st.divider()

#             # --- STEP 4: Process Downloads in Background ---
#             # Now that all charts are visible, we generate files in parallel
#             if download_queue:
#                 # Helper to process a single task
#                 def process_task(task):
#                     try:
#                         if task['type'] == 'excel':
#                             data = dataframe_to_formatted_excel(task['df'], task['lookback'], task['std_mult'])
#                             mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#                             ext = "xlsx"
#                         else:
#                             data = task['df'].to_csv(index=False).encode('utf-8')
#                             mime = "text/csv"
#                             ext = "csv"
#                         return task, data, mime, ext
#                     except Exception as e:
#                         return task, None, None, None

#                 # Run file generation in parallel
#                 with concurrent.futures.ThreadPoolExecutor() as executor:
#                     futures = [executor.submit(process_task, task) for task in download_queue]
                    
#                     for future in concurrent.futures.as_completed(futures):
#                         task, data, mime, ext = future.result()
#                         if data:
#                             # Update the specific placeholder associated with this chart
#                             task['placeholder'].download_button(
#                                 label=f"📊 Download {task['type'].upper()}",
#                                 data=data,
#                                 file_name=f"{task['filename']}.{ext}",
#                                 mime=mime,
#                                 key=task['filename']
#                             )
#                         else:
#                             task['placeholder'].error("Error generating file.")

#             else: 
#                 st.error("Not enough overlapping data found for this pair.")
#     else: 
#         st.warning("⚠️ Please select two **different** stocks in the sidebar.")



import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import concurrent.futures # For Parallel Processing

# Custom Modules
from modules.calculation import calculations
from modules.plotting import plotting
from modules import data_loader

# --- UI STYLING & CSS ---
def load_custom_css():
    st.markdown("""
        <style>
        /* Main Container Padding */
        .block-container {
            padding-top: 2rem;
            padding-bottom: 3rem;
        }
        
        /* Metrics Styling */
        div[data-testid="stMetricValue"] {
            font-size: 1.4rem;
            font-weight: 700;
            color: #ffffff;
        }
        div[data-testid="stMetricLabel"] {
            font-size: 0.9rem;
            color: #a0a0a0;
        }
        
        /* Custom Cards for Sections */
        .stContainer {
            background-color: #1e1e1e; /* Dark Card Background */
            border-radius: 10px;
        }
        
        /* Headers */
        h1, h2, h3 {
            font-family: 'Segoe UI', sans-serif;
            font-weight: 600;
        }
        h2 {
            color: #00e5ff; /* Cyan Highlight */
            border-bottom: 1px solid #333;
            padding-bottom: 10px;
            margin-top: 30px;
        }
        
        /* Tables */
        .stDataFrame {
            border: 1px solid #333;
            border-radius: 5px;
        }
        </style>
    """, unsafe_allow_html=True)

# --- HELPERS (Excel & Preprocess) ---
def dataframe_to_formatted_excel(df, lookback, std_mult):
    df_processed = df.copy()
    stock_a = str((df_processed['SYMBOL_A']).unique()[0])
    stock_b = str((df_processed['SYMBOL_B']).unique()[0])
    
    try:
        df_processed.drop(['INSTRUMENT_A','SYMBOL_A','OPEN_INT*_A','TRD_VAL_A','TRD_QTY_A','NO_OF_CONT_A','NO_OF_TRADE_A',
                           'TIMESTAMP_B','INSTRUMENT_B','SYMBOL_B','OPEN_INT*_B','TRD_VAL_B','TRD_QTY_B','NO_OF_CONT_B','NO_OF_TRADE_B',
                           'NEXT_EXP_DATE_A','NEXT_EXP_DATE_B'], axis=1, inplace=True)
    except KeyError: pass
    
    if 'TIMESTAMP_A' in df_processed.columns: 
        df_processed.rename(columns={'TIMESTAMP_A': 'Date'}, inplace=True)

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_processed.to_excel(writer, sheet_name="Pair Data", index=False, startrow=1)
        wb = writer.book
        ws = writer.sheets["Pair Data"]
        
        ws.insert_cols(14); ws.insert_cols(9); ws.insert_cols(2)
        
        # Headers
        ws.merge_cells("C1:H1"); ws["C1"] = stock_a
        ws["C1"].font = Font(bold=True); ws["C1"].alignment = Alignment(horizontal="center")
        
        ws.merge_cells("J1:O1"); ws["J1"] = stock_b
        ws["J1"].font = Font(bold=True); ws["J1"].alignment = Alignment(horizontal="center")
        
        ws.merge_cells("Q1:W1")
        ws["Q1"] = f"INDICATORS (RSI/Z-Score: {lookback}D | Bands: ±{std_mult} SD)"
        ws["Q1"].font = Font(bold=True); ws["Q1"].alignment = Alignment(horizontal="center")

        # Auto-width
        for col in ws.columns:
            max_length = 0
            column = col[1].column_letter 
            for cell in col:
                try: 
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = max(12, (max_length * 1.2))

    buffer.seek(0)
    return buffer

def preprocess_futures(df):
    df = df.copy()
    df.drop(['OPEN_INT_A', 'OPEN_INT_B'], axis=1, inplace=True, errors='ignore')
    for leg in ['A', 'B']:
        date_col = f'TIMESTAMP_{leg}'; expiry_col = f'EXP_DATE_{leg}'; symbol_col = f'SYMBOL_{leg}'; gap_col = f'expiry_gap_{leg}'
        df[date_col] = pd.to_datetime(df[date_col]); df[expiry_col] = pd.to_datetime(df[expiry_col])
        df = df.sort_values([symbol_col, expiry_col, date_col])
        next_exp_col = f'NEXT_EXP_DATE_{leg}'
        df[next_exp_col] = (df.groupby(symbol_col)[expiry_col].shift(-1))
        df[gap_col] = np.where(df[date_col] < df[expiry_col], (df[expiry_col] - df[date_col]).dt.days, (df[next_exp_col] - df[date_col]).dt.days)
        df.loc[df[next_exp_col].isna(), gap_col] = np.nan
    return df

# --- RENDERING FUNCTIONS ---
def render_chart_deferred(download_queue, fig, df_to_download, lookback, std_mult, filename_prefix, file_type="excel"):
    # 1. Show Chart Immediately
    st.plotly_chart(fig, use_container_width=True)
    
    # 2. Create Placeholder with nice styling
    col1, col2 = st.columns([0.8, 0.2])
    with col1:
        st.caption(f"ℹ️ Data Reference: {filename_prefix.replace('_', ' ').title()}")
    with col2:
        placeholder = st.empty()
        placeholder.caption(f"⏳ Gen...")

    # 3. Add to Queue
    download_queue.append({
        "placeholder": placeholder,
        "df": df_to_download,
        "lookback": lookback,
        "std_mult": std_mult,
        "filename": filename_prefix,
        "type": file_type
    })

# --- DATA CACHING ---
@st.cache_data(show_spinner=False)
def get_cached_data_and_calc(base_dir, sector, stock_a, stock_b, lookback, std_dev_mult, norm_start_date):
    # 1. Load
    raw_df = data_loader.load_pair_data(base_dir, sector, stock_a, stock_b)
    if raw_df is None or raw_df.empty:
        return None, None, None, None, None, None, None, None

    # 2. Preprocess
    raw_df = preprocess_futures(raw_df)

    # 3. Calculate
    processed_df = calculations.calculate_pair_metrics(raw_df, window=lookback, std_dev_multiplier=std_dev_mult)
    processed_df = calculations.calculate_quarterly_structure(processed_df)
    processed_df_for_semiannual = calculations.calculate_semiannual_weekly_structure(processed_df)
    
    stats = calculations.calculate_bollinger_reversion_stats(processed_df)
    spread = calculations.get_spread_series(processed_df).tail(250)
    
    rel_metrics = calculations.calculate_relationship_metrics(processed_df)
    coint_stats = calculations.calculate_cointegration_metrics(spread)
    hurst_value = calculations.calculate_hurst_exponent(spread)
    
    # 4. Normalize Data
    norm_df = calculations.calculate_normalized_performance(raw_df, norm_start_date)
    
    return (
        processed_df, stats, norm_df, raw_df, 
        processed_df_for_semiannual, coint_stats, 
        hurst_value, rel_metrics
    )

# --- PLOTTING ---
def generate_all_plots_parallel(processed_df, norm_df, stock_a, stock_b, std_dev_mult, norm_start_date_str, processed_df_for_semiannual):
    with concurrent.futures.ThreadPoolExecutor() as executor:
        future_ratio = executor.submit(plotting.plot_pair_ratio, processed_df, stock_a, stock_b, std_dev_mult)
        
        future_norm = None
        if not norm_df.empty:
            future_norm = executor.submit(plotting.plot_normalized_comparison, norm_df, stock_a, stock_b, norm_start_date_str)
            
        future_rsi = executor.submit(plotting.plot_secondary_indicators, processed_df, std_dev_mult)
        # Using the NEW plot_zscore_distribution function with zones
        future_dist = executor.submit(plotting.plot_zscore_distribution, processed_df) 
        future_structure = executor.submit(plotting.plot_quarterly_structure, processed_df, stock_a, stock_b)
        future_annual = executor.submit(plotting.plot_semiannual_weekly_structure, processed_df_for_semiannual, stock_a, stock_b)

        res_ratio = future_ratio.result()
        res_norm = future_norm.result() if future_norm else (None, None)
        res_rsi = future_rsi.result()
        res_dist = future_dist.result()
        res_structure = future_structure.result()
        res_annual = future_annual.result()
        
    return res_ratio, res_norm, res_rsi, res_dist, res_structure, res_annual


# --- MAIN RUN FUNCTION ---
def run(BASE_DIR, selected_sector, stock_a, stock_b, lookback, std_dev_mult, norm_start_date):
    
    load_custom_css()
    
    st.title("📈 Pro Quant Chart Studio")
    st.markdown("---")
    
    if not selected_sector:
        st.info("👈 Please select a sector from the sidebar to begin.")
        return

    col_info_1, col_info_2 = st.columns([0.7, 0.3])
    with col_info_1:
        st.markdown(f"### **{stock_a}** vs **{stock_b}**")
        st.caption(f"Sector: {selected_sector} | Strategy: Bollinger Mean Reversion")
    
    if stock_a != stock_b:
        # --- STEP 1: Fast Data & Calculation (Cached) ---
        with st.spinner("⚡ Crunching Financial Models..."):
            processed_df, stats, norm_df, raw_df, processed_df_for_semiannual, coint_stats, hurst, rel_metrics = get_cached_data_and_calc(
                BASE_DIR, selected_sector, stock_a, stock_b, lookback, std_dev_mult, norm_start_date
            )

        if processed_df is not None:
            
            # --- STEP 2: KPI DASHBOARD ---
            st.markdown("### 🚀 Strategy Performance")
            
            # Creating a Card-like container for metrics
            with st.container(border=True):
                m1, m2, m3, m4 = st.columns(4)
                m1.metric("-2σ Opportunities", stats["touches"], delta_color="normal")
                m2.metric("Mean Reversions", f"{stats['mean_reversions']}", delta=f"{stats['success_rate']:.1f}% Success")
                m3.metric("Failures", stats["failures"], delta_color="inverse")
                m4.metric("Avg Days to Profit", f"{stats['avg_days_to_mean']:.1f} Days")

            # Generate Plots in Background
            (fig_ratio, data_ratio), (fig_norm, data_norm), (fig_rsi, data_rsi), fig_dist, (fig_struct, data_struct), (fig_annual, plot_df_annual) = generate_all_plots_parallel(
                processed_df, norm_df, stock_a, stock_b, std_dev_mult, str(norm_start_date), processed_df_for_semiannual
            )

            # --- STEP 3: VISUALIZATIONS ---
            download_queue = []

            # --- CHART 1: MAIN RATIO ---
            st.markdown("## 1. Price Ratio & Bollinger Bands")
            with st.container(border=True):
                render_chart_deferred(download_queue, fig_ratio, data_ratio, lookback, std_dev_mult, "bollinger_ratio_data", "excel")
            
            
            # --- METRICS: RELATIONSHIP & STABILITY ---
            st.markdown("## 📊 Statistical Relationship Analysis")
            
            col_rel, col_risk = st.columns(2)
            
            with col_rel:
                with st.container(border=True):
                    st.subheader("Correlation & Spread")
                    r1, r2 = st.columns(2)
                    r1.metric("Correlation", f"{rel_metrics['correlation']:.2f}", delta="Strong" if rel_metrics['correlation'] > 0.6 else "Weak")
                    r2.metric("Spread Variance", f"{rel_metrics['spread_variance']:.4f}")
                    st.metric("Covariance", f"{rel_metrics['covariance']:.4f}")

            with col_risk:
                with st.container(border=True):
                    st.subheader("Stationarity Tests")
                    c1, c2 = st.columns(2)
                    c1.metric("Cointegration", "YES" if coint_stats["cointegrated"] else "NO", delta=f"Strength: {coint_stats['strength']:.2f}")
                    c2.metric("Hurst Exponent", f"{hurst:.2f}", delta="Mean Reverting" if hurst < 0.5 else "Trending", delta_color="inverse")
                    st.caption(f"ADF Stat: {coint_stats['adf_stat']:.2f} | p-value: {coint_stats['p_value']:.4f}")


            # --- CHART 4: REGIME ANALYSIS (Z-SCORE) ---
            st.markdown("## 2. Z-Score Regime Analysis")
            
            # Calculating distribution stats dynamically
            dist_stats = calculations.calculate_distribution_stats(processed_df)
            
            with st.container(border=True):
                col_graph, col_stats = st.columns([3, 1]) 
                
                with col_graph:
                    st.plotly_chart(fig_dist, use_container_width=True)
                
                with col_stats:
                    st.markdown("#### 📊 Sigma Counts")
                    if dist_stats is not None:
                        # Styling the DataFrame for a "Pro" look
                        st.dataframe(
                            dist_stats[['Range', 'Actual_%', 'Theoretical_%']].style.format({
                                'Actual_%': '{:.1f}%',
                                'Theoretical_%': '{:.1f}%'
                            }).background_gradient(subset=['Actual_%'], cmap='Greens'),
                            height=200,
                            use_container_width=True,
                            hide_index=True
                        )
                        
                        # Quick Insight
                        outlier_pct = dist_stats.iloc[3]['Actual_%']
                        if outlier_pct > 1.0:
                            st.error(f"⚠️ **High Risk:** {outlier_pct:.1f}% data in outliers (>3σ)!")
                        else:
                            st.success("✅ **Stable:** Distribution follows Normal Curve.")

            # --- CHART 2 & 3: NORMALIZED & RSI ---
            c_norm, c_rsi = st.columns(2)
            
            with c_norm:
                st.markdown("### Normalized Performance")
                with st.container(border=True):
                    if fig_norm:
                        render_chart_deferred(download_queue, fig_norm, data_norm, lookback, std_dev_mult, "normalized_data", "excel")
                    else:
                        st.warning("No data found for Normalized Chart.")

            with c_rsi:
                st.markdown("### RSI Momentum")
                with st.container(border=True):
                    render_chart_deferred(download_queue, fig_rsi, data_rsi, lookback, std_dev_mult, "rsi_data", "excel")
            
            
            # --- STRUCTURE ANALYSIS ---
            st.markdown("## 3. Structural Analysis (Quarterly & Annual)")
            
            tab1, tab2 = st.tabs(["📅 Quarterly Structure", "🗓️ Semi-Annual Structure"])
            
            with tab1:
                 render_chart_deferred(download_queue, fig_struct, data_struct, lookback, std_dev_mult, "quarterly_structure", "excel")
            
            with tab2:
                 render_chart_deferred(download_queue, fig_annual, plot_df_annual, lookback, std_dev_mult, "annualy_structure", "csv")
            
            st.divider()

            # --- STEP 4: BACKGROUND DOWNLOAD PROCESSING ---
            if download_queue:
                def process_task(task):
                    try:
                        if task['type'] == 'excel':
                            data = dataframe_to_formatted_excel(task['df'], task['lookback'], task['std_mult'])
                            mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            ext = "xlsx"
                        else:
                            data = task['df'].to_csv(index=False).encode('utf-8')
                            mime = "text/csv"
                            ext = "csv"
                        return task, data, mime, ext
                    except Exception as e:
                        return task, None, None, None

                with concurrent.futures.ThreadPoolExecutor() as executor:
                    futures = [executor.submit(process_task, task) for task in download_queue]
                    
                    for future in concurrent.futures.as_completed(futures):
                        task, data, mime, ext = future.result()
                        if data:
                            task['placeholder'].download_button(
                                label=f"📥 Download {task['type'].upper()}",
                                data=data,
                                file_name=f"{task['filename']}.{ext}",
                                mime=mime,
                                key=task['filename'],
                                use_container_width=True
                            )
                        else:
                            task['placeholder'].error("Failed")

        else: 
            st.error("📉 Not enough overlapping data found for this pair.")
    else: 
        st.warning("⚠️ Please select two **different** stocks in the sidebar.")